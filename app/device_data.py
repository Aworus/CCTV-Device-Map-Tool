from __future__ import annotations

import csv
import datetime as dt
import ipaddress
import json
import os
import re
import tempfile
import unicodedata
import zipfile
from collections import OrderedDict
from pathlib import Path
from xml.etree import ElementTree as ET

OFFICE_NS = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"
TABLE_NS = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
TEXT_NS = "urn:oasis:names:tc:opendocument:xmlns:text:1.0"
NS = {"office": OFFICE_NS, "table": TABLE_NS, "text": TEXT_NS}

SUPPORTED_SPREADSHEET_SUFFIXES = (".ods", ".xlsx", ".xls", ".csv")

# Opcjonalne, jawnie zatwierdzone rozstrzygnięcia dla zdublowanych wpisów IP.
# Wersja publiczna celowo nie zawiera wyjątków pochodzących z żadnego wdrożenia.
DUPLICATE_CAMERA_NAME_DECISIONS: dict[str, str] = {}

DEVICE_TYPES = ("camera", "recorder", "switch", "cabinet")
HOST_DEVICE_TYPES = ("camera", "recorder", "switch")
DEVICE_TYPE_LABELS = {
    "camera": "Kamery",
    "recorder": "Rejestratory",
    "switch": "Switche",
    "cabinet": "Szafy",
}
TYPE_DEFAULTS = {
    "camera": {
        "prefix": "CAM",
        "group": "CCTV / Kamery",
        "template": "ICMP Ping",
        "tag": "camera",
        "icon": "",
    },
    "recorder": {
        "prefix": "REC",
        "group": "CCTV / Rejestratory",
        "template": "ICMP Ping",
        "tag": "recorder",
        "icon": "RECORDER",
    },
    "switch": {
        "prefix": "SW",
        "group": "CCTV / Switche",
        "template": "ICMP Ping",
        "tag": "switch",
        "icon": "SWITCH",
    },
    # Szafa jest wyłącznie obiektem graficznym na mapie infrastruktury.
    # Nie reprezentuje hosta, dlatego nie ma IP, grupy, templateki ani pingu.
    "cabinet": {
        "prefix": "CAB",
        "group": "",
        "template": "",
        "tag": "",
        "icon": "CABINET",
    },
}

DEVICE_COLUMNS = [
    "Order",
    "Type",
    "IP",
    "HostName",
    "VisibleName",
    "GroupName",
    "TemplateName",
    "TagType",
    "IconName",
    # Zachowuje pochodzenie wpisu niezależnie od aktualnego źródła danych.
    # Dzięki temu ręcznie dodany host nadal może być bezpiecznie usunięty z
    # narzędzia po tym, jak jego dane zostaną potwierdzone przez Zabbixa.
    "Origin",
    "Source",
    "MapStatus",
    "X",
    "Y",
    "XPercent",
    "YPercent",
    "UpdatedAt",
    # Słownik JSON: nazwa mapy Zabbixa -> selementid dla szafy.
    # To jest wyłącznie lokalna metadana narzędzia. Dzięki niej ikona szafy
    # nie potrzebuje technicznego linku CCTV_TOOL_ID na mapie Zabbixa.
    "MapElementIds",
]


def normalize_device_type(value: str) -> str:
    text = str(value or "").strip().lower()
    aliases = {
        "camera": "camera",
        "kamera": "camera",
        "kamery": "camera",
        "recorder": "recorder",
        "rejestrator": "recorder",
        "rejestratory": "recorder",
        "switch": "switch",
        "switche": "switch",
        "przelacznik": "switch",
        "przełącznik": "switch",
        "cabinet": "cabinet",
        "szafa": "cabinet",
        "szafy": "cabinet",
    }
    if text not in aliases:
        raise ValueError(f"Nieznany typ urządzenia: {value!r}")
    return aliases[text]


def normalize_status(value: str) -> str:
    status = str(value or "PENDING").strip().upper()
    if status not in {"PLACED", "SKIPPED", "PENDING"}:
        return "PENDING"
    return status


def valid_ip(value: str) -> bool:
    try:
        ipaddress.ip_address(str(value).strip())
        return True
    except ValueError:
        return False


def technical_host_name(device_type: str, ip: str) -> str:
    device_type = normalize_device_type(device_type)
    prefix = TYPE_DEFAULTS[device_type]["prefix"]
    return f"{prefix}_{str(ip).strip().replace('.', '_')}"


def cabinet_local_id(existing_rows: list[dict], visible_name: str) -> str:
    """Tworzy stabilny identyfikator lokalnej szafy, bez hosta w Zabbixie."""
    normalized = unicodedata.normalize("NFKD", str(visible_name or ""))
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii").upper()
    slug = re.sub(r"[^A-Z0-9]+", "_", ascii_name).strip("_")[:32] or "SZAFA"
    base = f"CAB_{slug}"
    used = {str(row.get("HostName") or "").upper() for row in existing_rows}
    candidate = base
    suffix = 2
    while candidate.upper() in used:
        candidate = f"{base}_{suffix}"
        suffix += 1
    return candidate


def parse_map_element_ids(value: object) -> dict[str, str]:
    """Odczytuje lokalne powiązania szaf z elementami map Zabbixa.

    Dane są celowo przechowywane po stronie narzędzia, a nie jako link URL na
    mapie. Niepoprawna lub stara pusta wartość jest traktowana jak brak danych,
    aby aktualizacja była zgodna ze starszymi plikami devices.csv.
    """
    if isinstance(value, dict):
        source = value
    else:
        raw = str(value or "").strip()
        if not raw:
            return {}
        try:
            source = json.loads(raw)
        except (TypeError, ValueError, json.JSONDecodeError):
            return {}
    if not isinstance(source, dict):
        return {}
    return {
        str(map_name).strip(): str(element_id).strip()
        for map_name, element_id in source.items()
        if str(map_name).strip() and str(element_id).strip()
    }


def serialize_map_element_ids(value: object) -> str:
    """Zapisuje powiązania w stabilnej, kompaktowej postaci CSV."""
    return json.dumps(
        parse_map_element_ids(value),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def default_row(
    device_type: str,
    ip: str,
    host_name: str,
    visible_name: str,
    *,
    source: str,
    origin: str | None = None,
) -> dict:
    device_type = normalize_device_type(device_type)
    defaults = TYPE_DEFAULTS[device_type]
    source_text = str(source).strip()
    if origin is None:
        if device_type == "cabinet":
            origin = "CABINET"
        elif source_text.upper() == "ZABBIX":
            origin = "ZABBIX"
        elif source_text.upper() == "MANUAL":
            origin = "MANUAL"
        else:
            origin = "LOCAL"
    return {
        "Order": "0",
        "Type": device_type,
        "IP": str(ip).strip(),
        "HostName": str(host_name).strip(),
        "VisibleName": str(visible_name).strip(),
        "GroupName": defaults["group"],
        "TemplateName": defaults["template"],
        "TagType": defaults["tag"],
        "IconName": defaults["icon"],
        "Origin": str(origin).strip().upper() or "LOCAL",
        "Source": source_text,
        "MapStatus": "PENDING",
        "X": "",
        "Y": "",
        "XPercent": "",
        "YPercent": "",
        "UpdatedAt": "",
        "MapElementIds": "{}",
    }


def normalize_row(row: dict) -> dict:
    device_type = normalize_device_type(row.get("Type", "camera"))
    defaults = TYPE_DEFAULTS[device_type]
    source = str(row.get("Source") or "UNKNOWN").strip()
    legacy_source = source.upper()
    inferred_origin = (
        "CABINET"
        if device_type == "cabinet"
        else "MANUAL"
        if legacy_source == "MANUAL"
        else "ZABBIX"
        if legacy_source == "ZABBIX"
        else "LOCAL"
    )
    normalized = {
        "Order": str(row.get("Order") or "0").strip(),
        "Type": device_type,
        "IP": str(row.get("IP") or "").strip(),
        "HostName": str(row.get("HostName") or "").strip(),
        "VisibleName": str(row.get("VisibleName") or "").strip(),
        "GroupName": str(row.get("GroupName") or defaults["group"]).strip(),
        "TemplateName": str(row.get("TemplateName") or defaults["template"]).strip(),
        "TagType": str(row.get("TagType") or defaults["tag"]).strip(),
        "IconName": str(row.get("IconName") or defaults["icon"]).strip(),
        "Origin": str(row.get("Origin") or inferred_origin).strip().upper(),
        "Source": source,
        "MapStatus": normalize_status(row.get("MapStatus", "PENDING")),
        "X": str(row.get("X") or "").strip(),
        "Y": str(row.get("Y") or "").strip(),
        "XPercent": str(row.get("XPercent") or "").strip(),
        "YPercent": str(row.get("YPercent") or "").strip(),
        "UpdatedAt": str(row.get("UpdatedAt") or "").strip(),
        "MapElementIds": serialize_map_element_ids(row.get("MapElementIds")),
    }
    if not normalized["HostName"] and normalized["IP"]:
        normalized["HostName"] = technical_host_name(device_type, normalized["IP"])
    if not normalized["VisibleName"]:
        normalized["VisibleName"] = normalized["IP"] or normalized["HostName"]
    return normalized


def sort_devices(rows: list[dict]) -> list[dict]:
    order = {"camera": 0, "recorder": 1, "switch": 2, "cabinet": 3}

    def ip_sort(ip: str) -> tuple[int, int, int, int]:
        try:
            parts = tuple(int(part) for part in ip.split("."))
            if len(parts) == 4:
                return parts
        except ValueError:
            pass
        return (999, 999, 999, 999)

    normalized = [normalize_row(row) for row in rows]
    normalized.sort(
        key=lambda row: (
            order.get(row["Type"], 9),
            ip_sort(row["IP"]),
            row["HostName"].casefold(),
        )
    )
    counters = {item: 0 for item in DEVICE_TYPES}
    for row in normalized:
        counters[row["Type"]] += 1
        row["Order"] = str(counters[row["Type"]])
    return normalized


def read_devices(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter=";")
        fields = set(reader.fieldnames or [])
        required = {"Type", "IP", "HostName", "MapStatus"}
        if not required.issubset(fields):
            raise ValueError(
                "devices.csv nie ma wymaganych kolumn: Type, IP, HostName, MapStatus."
            )
        return sort_devices([normalize_row(row) for row in reader])


def write_devices(path: Path, rows: list[dict]) -> None:
    rows = sort_devices(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(
        prefix=path.stem + "_", suffix=".tmp", dir=str(path.parent)
    )
    os.close(descriptor)
    try:
        with Path(temp_name).open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=DEVICE_COLUMNS, delimiter=";")
            writer.writeheader()
            for row in rows:
                writer.writerow({column: row.get(column, "") for column in DEVICE_COLUMNS})
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def migrate_camera_points(camera_points_path: Path, devices_path: Path) -> dict:
    if devices_path.is_file():
        return {"created": False, "count": len(read_devices(devices_path))}
    rows: list[dict] = []
    if camera_points_path.is_file():
        with camera_points_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file, delimiter=";")
            fields = set(reader.fieldnames or [])
            if not {"IP", "HostName", "MapStatus"}.issubset(fields):
                raise ValueError("Stary camera_points.csv nie ma wymaganych kolumn.")
            for source in reader:
                row = default_row(
                    "camera",
                    source.get("IP", ""),
                    source.get("HostName", ""),
                    source.get("VisibleName", ""),
                    source="MIGRACJA_CAMERA_POINTS",
                )
                row.update(
                    {
                        "Order": source.get("Order", "0"),
                        "MapStatus": normalize_status(source.get("MapStatus", "PENDING")),
                        "X": source.get("X", ""),
                        "Y": source.get("Y", ""),
                        "XPercent": source.get("XPercent", ""),
                        "YPercent": source.get("YPercent", ""),
                        "UpdatedAt": source.get("UpdatedAt", ""),
                    }
                )
                rows.append(row)
    write_devices(devices_path, rows)
    return {"created": True, "count": len(rows)}


def upsert_devices(
    existing_rows: list[dict],
    incoming_rows: list[dict],
    *,
    preserve_ods_camera_names: bool = False,
) -> dict:
    existing_rows = [normalize_row(row) for row in existing_rows]
    incoming_rows = [normalize_row(row) for row in incoming_rows]
    by_host = {(row["Type"], row["HostName"]): row for row in existing_rows}
    by_ip = {(row["Type"], row["IP"]): row for row in existing_rows if row["IP"]}
    added = 0
    updated = 0
    protected = {
        "Origin",
        "MapStatus",
        "X",
        "Y",
        "XPercent",
        "YPercent",
        "UpdatedAt",
    }

    for incoming in incoming_rows:
        current = by_host.get((incoming["Type"], incoming["HostName"]))
        if current is None and incoming["IP"]:
            current = by_ip.get((incoming["Type"], incoming["IP"]))
        if current is None:
            existing_rows.append(incoming)
            by_host[(incoming["Type"], incoming["HostName"])] = incoming
            if incoming["IP"]:
                by_ip[(incoming["Type"], incoming["IP"])] = incoming
            added += 1
            continue
        before = dict(current)
        # Nazwa oczekująca na wysłanie (z arkusza albo ręcznej edycji) ma
        # pierwszeństwo przed odczytem startowym z Zabbixa. Dzięki temu
        # odświeżenie listy nie „zgubi” zmiany, której operator jeszcze nie
        # zatwierdził przyciskiem aktualizacji nazw.
        pending_source = str(current.get("Source") or "").upper()
        local_pending_name = (
            preserve_ods_camera_names
            and (
                pending_source.startswith("ARKUSZ — OCZEKUJE")
                or pending_source.startswith("EDYCJA LOKALNA — OCZEKUJE")
            )
        )
        for column in DEVICE_COLUMNS:
            if column in protected:
                continue
            value = incoming.get(column, "")
            if value == "":
                continue
            if local_pending_name and column in {"VisibleName", "Source"}:
                continue
            current[column] = value
        if current != before:
            updated += 1

    return {"rows": sort_devices(existing_rows), "added": added, "updated": updated}


def summarize_devices(rows: list[dict]) -> dict:
    result = {}
    for device_type in DEVICE_TYPES:
        selected = [row for row in rows if row["Type"] == device_type]
        counts = {
            "total": len(selected),
            "PLACED": 0,
            "SKIPPED": 0,
            "PENDING": 0,
            "duplicates": 0,
        }
        seen_hosts: set[str] = set()
        seen_ips: set[str] = set()
        for row in selected:
            counts[normalize_status(row["MapStatus"])] += 1
            host = row["HostName"]
            ip = row["IP"]
            if (host and host in seen_hosts) or (ip and ip in seen_ips):
                counts["duplicates"] += 1
            if host:
                seen_hosts.add(host)
            if ip:
                seen_ips.add(ip)
        result[device_type] = counts
    return result


def parse_optional_float(value: str | None) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def short_device_label(row: dict) -> str:
    if row.get("Type") == "cabinet":
        return str(row.get("VisibleName") or "Szafa")[:18]
    ip = str(row.get("IP") or "").strip()
    if valid_ip(ip):
        parts = ip.split(".")
        return parts[-1] if row.get("Type") == "camera" else f"{parts[-2]}.{parts[-1]}"
    return str(row.get("VisibleName") or "?")[:12]


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower().replace("\n", " "))


def cell_text(cell: ET.Element) -> str:
    return "\n".join(
        "".join(paragraph.itertext())
        for paragraph in cell.findall(".//text:p", NS)
    ).strip()


def parse_physical_ods_rows(sheet: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in sheet.findall("table:table-row", NS):
        values: list[str] = []
        for cell in list(row):
            if cell.tag not in {
                f"{{{TABLE_NS}}}table-cell",
                f"{{{TABLE_NS}}}covered-table-cell",
            }:
                continue
            repeat = int(cell.attrib.get(f"{{{TABLE_NS}}}number-columns-repeated", "1"))
            value = cell_text(cell) if cell.tag == f"{{{TABLE_NS}}}table-cell" else ""
            if value == "" and repeat > 1000:
                break
            values.extend([value] * min(repeat, 1000))
        while values and values[-1] == "":
            values.pop()
        if any(value.strip() for value in values):
            rows.append(values)
    return rows


def short_camera_name(ip: str) -> str:
    """Zwraca skróconą, dwuczłonową końcówkę prawidłowego adresu IP."""
    parts = str(ip).strip().split(".")
    return f"{parts[-2]}.{parts[-1]}" if len(parts) == 4 else str(ip).strip()


def camera_name_key(value: str) -> str:
    """Porównanie nazw bez wielkości liter, spacji i polskich znaków."""
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_value.casefold())


def resolve_camera_ip_from_sheet(name: str, listed_ip: str) -> tuple[str, bool]:
    """Ustala adres kamery, gdy w kolumnie ``Adres IP`` jest skopiowany IP
    sąsiedniej kamery.

    W części spisów rejestrator zapisuje ten sam adres w kolumnie G dla kilku
    kanałów, ale ich rzeczywiste adresy nadal są jednoznacznie zapisane na
    początku nazwy kamery (np. ``20.25 Opis``).
    W takim przypadku F jest bezpieczniejszym identyfikatorem niż G. Zmieniamy
    tylko ostatni oktet, zachowując sieć z G; w każdym innym przypadku nadal
    używamy wartości z G.
    """
    source_ip = str(listed_ip or "").strip()
    if not valid_ip(source_ip):
        return source_ip, False

    normalized_name = " ".join(str(name or "").replace("\n", " ").split())
    source_parts = source_ip.split(".")

    # Obsługa pełnego adresu wpisanego na początku F.
    full_match = re.match(r"^((?:\d{1,3}\.){3}\d{1,3})(?=\D|$)", normalized_name)
    if full_match:
        candidate = full_match.group(1)
        if valid_ip(candidate) and candidate.split(".")[:3] == source_parts[:3]:
            return candidate, candidate != source_ip

    # Standardowy format w arkuszu: ``222.221 Opis`` albo ``222.173WSP``.
    short_match = re.match(r"^(\d{1,3}\.\d{1,3})(?=\D|$)", normalized_name)
    if short_match:
        short_parts = short_match.group(1).split(".")
        if short_parts[0] == source_parts[2]:
            candidate = ".".join([source_parts[0], source_parts[1], *short_parts])
            if valid_ip(candidate):
                return candidate, candidate != source_ip

    return source_ip, False


def camera_visible_name_from_sheet(name: str, info: str, ip: str) -> tuple[str, bool]:
    """Wyznacza nazwę Zabbixa z faktycznej nazwy kamery (F).

    F jest źródłem prawdy dopiero wtedy, gdy zawiera końcówkę IP oraz opis
    zgodny z J. Gdy kamera jest niedostępna, F zwykle ma sam numer albo
    techniczną nazwę typu IPCamera — wtedy w Zabbixie zostaje tylko 222.xxx.
    Druga wartość mówi, czy nazwa z F została potwierdzona przez J.
    """
    actual_name = " ".join(str(name or "").replace("\n", " ").split())
    expected_info = " ".join(str(info or "").replace("\n", " ").split())
    short_name = short_camera_name(ip)

    remainder = ""
    if actual_name.casefold().startswith(short_name.casefold()):
        remainder = actual_name[len(short_name) :].strip()

    confirmed = bool(
        expected_info
        and expected_info.casefold() != "do ustalenia"
        and remainder
        and camera_name_key(remainder) == camera_name_key(expected_info)
    )
    return (actual_name if confirmed else short_name), confirmed


def clean_visible_name(name: str, info: str, ip: str) -> str:
    """Zgodność ze starszym kodem — zwraca tylko wyznaczoną nazwę."""
    return camera_visible_name_from_sheet(name, info, ip)[0]


def stringify_sheet_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat(sep=" ") if isinstance(value, dt.datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_ods_sheets(path: Path) -> list[list[list[str]]]:
    with zipfile.ZipFile(path) as archive:
        xml_data = archive.read("content.xml")
    root = ET.fromstring(xml_data)
    return [
        rows
        for sheet in root.findall(".//table:table", NS)
        if (rows := parse_physical_ods_rows(sheet))
    ]


def read_xlsx_sheets(path: Path) -> list[list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Brakuje obsługi XLSX (openpyxl). Uruchom instalator aktualizacji V2.6."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[list[list[str]]] = []
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for source_row in sheet.iter_rows(values_only=True):
                values = [stringify_sheet_value(value) for value in source_row]
                while values and values[-1] == "":
                    values.pop()
                if any(value.strip() for value in values):
                    rows.append(values)
            if rows:
                sheets.append(rows)
        return sheets
    finally:
        workbook.close()


def read_xls_sheets(path: Path) -> list[list[list[str]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "Brakuje obsługi starego XLS (xlrd). Uruchom instalator aktualizacji V2.6."
        ) from exc

    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        sheets: list[list[list[str]]] = []
        for sheet in workbook.sheets():
            rows: list[list[str]] = []
            for row_index in range(sheet.nrows):
                values = [
                    stringify_sheet_value(sheet.cell_value(row_index, column_index))
                    for column_index in range(sheet.ncols)
                ]
                while values and values[-1] == "":
                    values.pop()
                if any(value.strip() for value in values):
                    rows.append(values)
            if rows:
                sheets.append(rows)
        return sheets
    finally:
        workbook.release_resources()


def read_csv_sheets(path: Path) -> list[list[list[str]]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "cp1250", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Nie udało się rozpoznać kodowania pliku CSV.")

    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    rows: list[list[str]] = []
    for source_row in csv.reader(text.splitlines(), delimiter=delimiter):
        values = [str(value).strip() for value in source_row]
        while values and values[-1] == "":
            values.pop()
        if any(value.strip() for value in values):
            rows.append(values)
    return [rows] if rows else []


def read_spreadsheet_sheets(path: Path) -> list[list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SPREADSHEET_SUFFIXES:
        supported = ", ".join(SUPPORTED_SPREADSHEET_SUFFIXES)
        raise ValueError(f"Nieobsługiwany format {suffix or '(brak)'}. Obsługiwane: {supported}.")
    if suffix == ".ods":
        return read_ods_sheets(path)
    if suffix == ".xlsx":
        return read_xlsx_sheets(path)
    if suffix == ".xls":
        return read_xls_sheets(path)
    return read_csv_sheets(path)


def find_camera_table(sheets: list[list[list[str]]]) -> list[list[str]]:
    for rows in sheets:
        for header_row_index, row in enumerate(rows[:100]):
            headers = [normalize_header(value) for value in row]
            if "adres ip" in headers and "nazwa kamery" in headers:
                return [row, *rows[header_row_index + 1 :]]
    raise ValueError(
        "Nie znaleziono arkusza z kolumnami Adres IP i Nazwa Kamery."
    )


def read_cameras_from_spreadsheet(
    path: Path, *, return_diagnostics: bool = False
) -> list[dict] | tuple[list[dict], dict]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"Nie znaleziono pliku: {path}")
    selected_rows = find_camera_table(read_spreadsheet_sheets(path))

    headers = [normalize_header(value) for value in selected_rows[0]]
    header_index = {header: index for index, header in enumerate(headers) if header}

    def index_of(*names: str) -> int | None:
        for name in names:
            normalized = normalize_header(name)
            if normalized in header_index:
                return header_index[normalized]
        return None

    ip_index = index_of("Adres IP")
    name_index = index_of("Nazwa Kamery")
    info_index = index_of("222.X - [info]", "222.X - Info")
    if ip_index is None or name_index is None:
        raise ValueError("Brakuje wymaganych kolumn kamer w arkuszu.")

    candidates_by_ip: OrderedDict[str, list[dict]] = OrderedDict()
    # Wpisy, dla których kolumna G została zastąpiona numerem rozpoznanym
    # z początku F. Zachowujemy je do raportu, aby operator widział dokładnie
    # dlaczego importer nie zaufał skopiowanemu adresowi.
    ip_recovered_from_name: list[dict] = []
    raw_headers = selected_rows[0]
    for values in selected_rows[1:]:
        padded = values + [""] * max(0, len(raw_headers) - len(values))
        listed_ip = padded[ip_index].strip()
        if not listed_ip or not valid_ip(listed_ip):
            continue
        name = padded[name_index].strip()
        info = padded[info_index].strip() if info_index is not None else ""
        ip, recovered = resolve_camera_ip_from_sheet(name, listed_ip)
        if recovered:
            ip_recovered_from_name.append(
                {
                    "listed_ip": listed_ip,
                    "resolved_ip": ip,
                    "name": " ".join(name.replace("\n", " ").split()),
                }
            )
        visible_name, confirmed = camera_visible_name_from_sheet(name, info, ip)
        candidates_by_ip.setdefault(ip, []).append(
            {
                "name": " ".join(name.replace("\n", " ").split()),
                "info": " ".join(info.replace("\n", " ").split()),
                "visible_name": visible_name,
                "confirmed": confirmed,
            }
        )

    cameras: list[dict] = []
    diagnostics = {
        "confirmed_names": 0,
        "short_names": 0,
        "duplicate_ips": [],
        "ambiguous_name_ips": [],
        "ip_recovered_from_name": ip_recovered_from_name,
        "warnings": [],
    }
    for ip, candidates in candidates_by_ip.items():
        confirmed_names = []
        for candidate in candidates:
            value = candidate["visible_name"]
            if candidate["confirmed"] and value not in confirmed_names:
                confirmed_names.append(value)

        if len(confirmed_names) > 1:
            diagnostics["ambiguous_name_ips"].append(ip)

        decided_name = DUPLICATE_CAMERA_NAME_DECISIONS.get(ip)
        if decided_name and decided_name in confirmed_names:
            # Ręcznie zatwierdzona decyzja ma pierwszeństwo nad błędnym,
            # zdublowanym wpisem w arkuszu. Nie korzystamy z niej, jeśli
            # brak już potwierdzenia w F/J.
            visible_name = decided_name
            diagnostics["confirmed_names"] += 1
            diagnostics["warnings"].append(
                f"{ip}: użyto zatwierdzonej nazwy {decided_name!r} "
                "dla zdublowanego wpisu IP."
            )
        elif len(confirmed_names) == 1:
            visible_name = confirmed_names[0]
            diagnostics["confirmed_names"] += 1
        elif len(confirmed_names) > 1:
            # Dwie różne potwierdzone nazwy dla jednego IP są niejednoznaczne.
            # Nie zgadujemy — używamy bezpiecznej krótkiej nazwy i zapisujemy alert.
            visible_name = short_camera_name(ip)
            diagnostics["short_names"] += 1
            diagnostics["warnings"].append(
                f"{ip}: więcej niż jedna potwierdzona nazwa w kolumnach F/J. "
                "Użyto krótkiej nazwy IP."
            )
        else:
            visible_name = short_camera_name(ip)
            diagnostics["short_names"] += 1

        row = default_row(
            "camera",
            ip,
            technical_host_name("camera", ip),
            visible_name,
            source=path.suffix.lstrip(".").upper(),
        )
        cameras.append(row)

        if len(candidates) > 1:
            diagnostics["duplicate_ips"].append(ip)
            info_values = {candidate["info"] for candidate in candidates}
            if len(info_values) > 1:
                diagnostics["warnings"].append(
                    f"{ip}: zdublowany IP ma różne wartości J "
                    f"({', '.join(sorted(repr(value) for value in info_values))})."
                )

    if not cameras:
        raise ValueError("W arkuszu nie znaleziono poprawnych adresów kamer.")
    for item in ip_recovered_from_name:
        diagnostics["warnings"].append(
            "IP z kolumny F ma pierwszeństwo: "
            f"{item['listed_ip']} -> {item['resolved_ip']} "
            f"({item['name']!r})."
        )
    sorted_cameras = sort_devices(cameras)
    return (sorted_cameras, diagnostics) if return_diagnostics else sorted_cameras


def read_cameras_from_ods(path: Path) -> list[dict]:
    """Zgodność ze starszym kodem; nowy importer rozpoznaje format po rozszerzeniu."""
    return read_cameras_from_spreadsheet(path)
